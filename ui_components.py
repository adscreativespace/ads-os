"""
ADS OS Desktop -- Shared UI Components
Reusable widgets so future screens are assembled from these instead of each
file writing its own near-identical stat-card/badge/table code. Dashboard is
the first consumer (dogfooding this immediately, not building speculative
components nothing calls). Migrating the other modules' existing local card-
rendering code (Materials, Vendors, BOQ, Invoice Center, Commercial Reports)
to use these is a separate, lower-risk, mechanical future pass -- not
bundled into this change to keep blast radius contained and testable.
"""
import customtkinter as ctk
import tkinter
from tkinter import ttk
import datetime
import theme

ACTIVITY_ICONS = {
    "Created": "✚", "Updated": "✎", "Deleted": "🗑", "Report Generated": "📄",
    "Proposal PDF Generated": "📄", "Proposal Created": "📝", "Payment Recorded": "💳",
    "Invoice Created": "🧾", "Fee Calculation Saved": "🧮",
    "Project Opened": "📂", "Contract Created": "📋", "Contract Payment Recorded": "💳",
    "Commission Receipt Recorded": "💰", "Contractor Added": "👷", "Contractor Archived": "🗄",
    "Contractor Updated": "✎", "Vendor Added": "🧱", "Vendor Archived": "🗄", "Vendor Updated": "✎",
    "Invoice Cancelled": "✕", "Invoice Marked Sent": "📤", "Invoice Updated": "✎",
    "Material Added": "📦", "Material Deleted": "🗑", "Material Purchase Recorded": "🛒",
    "Material Updated": "✎", "Purchase Recorded": "🛒", "Quotation Created": "📝",
    "BOQ Item Added": "📐", "BOQ Item Deleted": "🗑", "BOQ Item Updated": "✎",
}
DEFAULT_ACTIVITY_ICON = "•"

# Color grouped by what the action actually represents -- not a fixed
# per-string mapping (that would need updating every time a new Action
# string is added anywhere in the app), so any current or future Added/
# Created/Recorded-style action gets a sensible color automatically.
_ACTIVITY_COLOR_RULES = [
    (("Deleted", "Archived", "Cancelled"), "#8B2E2E"),   # removal-type actions -- muted red
    (("Added", "Created", "Recorded"), "#2E8B57"),         # creation/recording -- green
    (("Generated", "Sent"), "#5B4A8A"),  # document/output actions -- purple
    (("Updated",), "#3B6EA5"),                              # edits -- blue
]


def _activity_color(action):
    for keywords, color in _ACTIVITY_COLOR_RULES:
        if any(kw in action for kw in keywords):
            return color
    return theme.BRASS


def metric_row(parent, value, label, value_color=None):
    """
    A single big-value-over-small-label row, e.g. Financial Snapshot's
    '₹200,000 / Invoiced' style -- the one genuinely new, reusable pattern
    from Dashboard's polish pass that didn't already exist under a
    different name (stat_card and activity_card already covered the other
    two). Deliberately simpler than stat_card: no icon, no card background,
    no fixed width -- just a value+label pair meant to stack vertically
    inside an existing panel.
    """
    row = ctk.CTkFrame(parent, fg_color="transparent")
    row.pack(fill="x", padx=15, pady=4)
    ctk.CTkLabel(row, text=value, font=("Georgia", 17, "bold"), text_color=value_color or theme.BRASS,
                anchor="w").pack(fill="x", anchor="w")
    ctk.CTkLabel(row, text=label, font=("Segoe UI", 9), text_color=theme.MUTED, anchor="w").pack(fill="x", anchor="w")
    return row


def kpi_card(parent, value, label, sublabel, icon, bg_color, col_idx, value_size=27, on_click=None):
    """
    Shared KPI card used by Clients, Projects, Vendors, Contractors, and
    Financial Dashboard -- icon in a colored circle, bold title, large
    value, muted sublabel.

    Height=104 and title wraplength=115 (both increased from the previous
    90/no-wrap) fix a real, larger clipping bug found via "Contractor
    Outstanding" (22 chars, needing 160px against only 121px available) --
    font-size reduction alone (the previous fix, for a smaller ~7px
    shortfall) can't safely close a ~40px gap without risking the same
    illegibility mistake made once already. Verified via PIL against every
    real label in use across all five modules before choosing 115px: 6 of
    27 real labels (Contractor Outstanding, Total Contract Value, Vendor
    Outstanding, Client Receivables, Active Contractors, Total Contractors)
    genuinely need to wrap to 2 lines at this width; the other 21 stay on
    one line. This is an honest, width-based outcome, not a special case
    for one label -- any future label this long will also wrap correctly
    rather than clip, without needing another manual re-measurement pass.

    Optional value_size (default 27, matching every existing call site
    unchanged) lets one caller give a single card slightly more emphasis
    (e.g. Financial Dashboard's Net Position) without touching the shared
    styling every other card still uses -- purely additive, not a
    restyle of the component itself.

    Optional on_click makes the whole card clickable -- bound recursively
    to the card frame and every child label/icon, since Tkinter does not
    propagate click events from children up to a parent frame on its own.
    Cards without on_click behave exactly as before (no cursor change, no
    binding), so every existing non-clickable caller is unaffected.

    Caller is responsible for grid_columnconfigure(col, weight=1) on its
    own stats_frame for the responsive, evenly-gutters distribution --
    this function only builds and grids the individual card.
    """
    card = ctk.CTkFrame(parent, fg_color=theme.WHITE, corner_radius=10, width=175, height=104,
                       border_width=1, border_color="#E8E0D0")
    card.grid(row=0, column=col_idx, pady=(0, 0))
    card.pack_propagate(False)
    top_row = ctk.CTkFrame(card, fg_color="transparent")
    top_row.pack(fill="x", padx=12, pady=(10, 0), anchor="w")
    icon_bg = ctk.CTkLabel(top_row, text=icon, font=("Segoe UI", 12), fg_color=bg_color,
                           corner_radius=12, width=24, height=24)
    icon_bg.pack(side="left", anchor="n")
    title_label = ctk.CTkLabel(top_row, text=label, font=("Segoe UI", 12, "bold"), text_color=theme.MUTED,
                               wraplength=115, justify="left")
    title_label.pack(side="left", padx=(6, 0))
    value_label = ctk.CTkLabel(card, text=value, font=("Georgia", value_size, "bold"), text_color=theme.INK,
                               anchor="w")
    value_label.pack(anchor="w", padx=12, pady=(2, 0))
    sublabel_label = ctk.CTkLabel(card, text=sublabel, font=("Segoe UI", 10), text_color=theme.MUTED, anchor="w")
    sublabel_label.pack(anchor="w", padx=12, pady=(3, 0))

    if on_click:
        for widget in (card, top_row, icon_bg, title_label, value_label, sublabel_label):
            widget.configure(cursor="hand2") if hasattr(widget, "configure") else None
            widget.bind("<Button-1>", lambda e: on_click())
    return card


def stat_card(parent, value, label, sublabel=None, width=175, height=90, icon=None, sublabel_wrap=None):
    """
    A compact KPI card. Order is deliberately Icon -> small Title -> large
    Value -> small Context, not Icon -> Value -> Title -- the eye should
    land on the number after already knowing what it's counting, matching
    how a person actually reads a stat (title orients, value informs).
    `icon` remains purely additive; every existing call site across the
    app continues to render correctly without passing it.

    `sublabel_wrap` (default None, matching every existing call site
    exactly) lets a caller with a genuinely longer sublabel wrap it
    instead of risking a clip -- learned directly from kpi_card's earlier
    clipping bug rather than assume a new longer sublabel will just fit.
    """
    card = ctk.CTkFrame(parent, fg_color=theme.WHITE, corner_radius=8, width=width, height=height)
    card.pack_propagate(False)
    if icon:
        ctk.CTkLabel(card, text=icon, font=("Segoe UI", 15)).pack(pady=(10, 0))
    ctk.CTkLabel(card, text=label, font=theme.FONT_SMALL, text_color=theme.MUTED).pack(pady=(4 if icon else 12, 0))
    ctk.CTkLabel(card, text=str(value), font=("Georgia", 22, "bold"), text_color=theme.BRASS).pack()
    if sublabel:
        # wraplength must never be None here -- CTkLabel's own default is
        # 0 (an int), and passing None explicitly breaks its internal
        # scaling math (int(None * float)), crashing the instant this
        # renders. Only set wraplength when a real value was actually
        # requested; every other call omits it entirely, exactly matching
        # the widget's own safe default.
        label_kwargs = {"wraplength": sublabel_wrap} if sublabel_wrap else {}
        ctk.CTkLabel(card, text=sublabel, font=("Segoe UI", 9), text_color=theme.MUTED,
                    justify="center", **label_kwargs).pack()
    return card


def pill_badge(parent, text, color):
    """
    A filled, rounded chip badge -- e.g. '[ High Priority ]' with a solid
    color background, as opposed to status_badge's dot+text style. Added
    as a new function rather than changing status_badge in place, since
    that's also used by Projects and this pass is scoped to Clients.
    """
    return ctk.CTkLabel(parent, text=f"  {text}  ", font=theme.FONT_SMALL, text_color=theme.WHITE,
                        fg_color=color, corner_radius=10)


def status_badge(parent, text, color, font=None):
    """A small colored-dot + label badge, e.g. '● Design'."""
    row = ctk.CTkFrame(parent, fg_color="transparent")
    ctk.CTkLabel(row, text="●", font=("Segoe UI", 12), text_color=color).pack(side="left")
    ctk.CTkLabel(row, text=f" {text}", font=(font or theme.FONT_SMALL), text_color=theme.INK).pack(side="left")
    return row


def section_header(parent, title, action_text=None, action_command=None):
    """A card/section title row, optionally with a right-aligned action button."""
    row = ctk.CTkFrame(parent, fg_color="transparent")
    ctk.CTkLabel(row, text=title, font=theme.FONT_BODY_BOLD, text_color=theme.INK).pack(side="left")
    if action_text and action_command:
        ctk.CTkButton(row, text=action_text, command=action_command,
                      fg_color=theme.MUTED, hover_color=theme.INK, font=theme.FONT_SMALL,
                      width=80, height=24).pack(side="right")
    return row


def empty_state(parent, message, action_text=None, action_command=None):
    """
    Shown in place of a table/list when there's genuinely no data, instead
    of a silent blank table. `parent` should be the same container the
    Treeview lives in; call this after clearing the tree when the result
    set is empty, and destroy the returned frame (or just clear `parent`
    of non-tree children) before the next refresh.
    """
    frame = ctk.CTkFrame(parent, fg_color="transparent")
    frame.pack(fill="both", expand=True, pady=30)
    ctk.CTkLabel(frame, text=message, font=theme.FONT_BODY, text_color=theme.MUTED).pack()
    if action_text and action_command:
        ctk.CTkButton(frame, text=action_text, command=action_command, fg_color=theme.BRASS,
                     hover_color=theme.INK, font=theme.FONT_SMALL, height=28).pack(pady=(10, 0))
    return frame


def relative_time(timestamp_str):
    """Formats a stored timestamp (e.g. from logActivity) as 'Just now' / '10 min ago' / 'Today' / 'Yesterday' / a date."""
    try:
        ts = datetime.datetime.fromisoformat(timestamp_str)
    except (ValueError, TypeError):
        return timestamp_str or "-"
    now = datetime.datetime.now()
    delta = now - ts
    if delta.total_seconds() < 60:
        return "Just now"
    if delta.total_seconds() < 3600:
        mins = int(delta.total_seconds() // 60)
        return f"{mins} min ago"
    if ts.date() == now.date():
        return f"Today, {ts.strftime('%I:%M %p')}"
    if ts.date() == (now - datetime.timedelta(days=1)).date():
        return f"Yesterday, {ts.strftime('%I:%M %p')}"
    return ts.strftime("%d %b %Y")


def activity_card(parent, action, details, when):
    """One Recent Activity entry, styled as a small timeline row instead of a raw table row."""
    row = ctk.CTkFrame(parent, fg_color="transparent")
    row.pack(fill="x", pady=4)
    icon = ACTIVITY_ICONS.get(action, DEFAULT_ACTIVITY_ICON)
    color = _activity_color(action)
    ctk.CTkLabel(row, text=icon, font=("Segoe UI", 14), text_color=color, width=26).pack(side="left")
    text_frame = ctk.CTkFrame(row, fg_color="transparent")
    text_frame.pack(side="left", fill="x", expand=True)
    ctk.CTkLabel(text_frame, text=action, font=theme.FONT_BODY_BOLD,
                text_color=theme.INK, anchor="w").pack(fill="x")
    if details:
        ctk.CTkLabel(text_frame, text=details, font=("Segoe UI", 9), text_color=theme.MUTED, anchor="w").pack(fill="x")
    ctk.CTkLabel(row, text=relative_time(when), font=("Segoe UI", 9), text_color=theme.MUTED).pack(side="right")
    return row


class AdaptiveScrollFrame(ctk.CTkFrame):
    """
    A scrollable container whose scrollbars only appear when content
    genuinely overflows the visible area -- unlike CTkScrollableFrame,
    which always shows a scrollbar regardless of whether one is needed.
    Real requirement: comfortable side-by-side working (e.g. ADS OS resized
    next to WhatsApp while copying client data), where a permanently visible
    scrollbar wastes space and looks wrong when nothing actually needs
    scrolling.

    Usage: create one of these where you'd normally use
    ctk.CTkScrollableFrame, then build content inside `.content` instead of
    inside the frame directly.

        container = AdaptiveScrollFrame(parent)
        container.pack(fill="both", expand=True)
        ctk.CTkLabel(container.content, text="...").pack(...)

    Standard Tkinter canvas+scrollbar pattern -- not a CustomTkinter
    built-in, since CTkScrollableFrame doesn't support this. Both scrollbars
    are independent (vertical shows when content is taller than the
    viewport, horizontal shows when wider), matching "show only the one
    that's actually needed."
    """
    def __init__(self, master, fg_color="transparent", **kwargs):
        super().__init__(master, fg_color=fg_color, **kwargs)

        canvas_bg = fg_color if fg_color != "transparent" else theme.PARCHMENT
        self.canvas = tkinter.Canvas(self, highlightthickness=0, bg=canvas_bg)
        # Real fix for a real, reported bug: mouse wheel scrolling felt
        # sluggish across every screen using this shared component (not
        # just Clients -- this affects all of them equally, since the
        # cause was never screen-specific). yscrollincrement was never
        # set, defaulting to 0; with that default, _on_mousewheel's
        # yview_scroll(n, "units") call below moves the view by only a
        # tiny fraction of a pixel's worth per wheel notch, so a normal
        # scroll through a card grid or a table needs far more notches
        # than it should. First attempt used 90px per notch; reported
        # directly as still too slow, so increased further to 180px --
        # a noticeably bigger jump per notch, closer to how a fast native
        # scroll feels, without being so aggressive that a short list
        # overshoots past its own end in a single notch.
        self.canvas.configure(yscrollincrement=180, xscrollincrement=180)
        self.vbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.hbar = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.vbar.set, xscrollcommand=self.hbar.set)

        # Content lives in here, not directly in the canvas -- this is what
        # the caller actually builds widgets into.
        self.content = ctk.CTkFrame(self.canvas, fg_color="transparent")
        self._content_window = self.canvas.create_window((0, 0), window=self.content, anchor="nw")

        self.canvas.pack(side="left", fill="both", expand=True)
        # Scrollbars are NOT packed here -- _update_scrollbars() shows/hides
        # them on demand. Starting hidden matches "no scrollbars when
        # everything fits."

        self.content.bind("<Configure>", self._on_content_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.bind("<Configure>", lambda e: self._schedule_update())
        self._bind_mousewheel()

        self._debounce_id = None
        # Immediate check for the initial build, plus a debounced settle
        # check queued right behind it.
        self._update_scrollbars()
        self._schedule_update()

    def _schedule_update(self):
        """
        Debounced re-check: a window drag or fast sequence of content
        changes fires many Configure events in quick succession. Measuring
        on every single one risks measuring an intermediate, not-yet-final
        size -- if that intermediate measurement happens to be the LAST one
        that runs before events stop, the scrollbar state gets stuck
        reflecting a size that no longer exists. Cancelling any pending
        check and rescheduling on every event means only the state after
        things actually stop changing gets measured, not whatever the
        last event during the drag happened to be.
        """
        if self._debounce_id is not None:
            self.after_cancel(self._debounce_id)
        self._debounce_id = self.after(120, self._update_scrollbars)

    def _on_content_configure(self, _event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self._schedule_update()

    def _on_canvas_configure(self, event):
        # Keep the content at least as wide as the visible canvas, so short
        # content doesn't look left-aligned in a mostly-empty canvas --
        # matches how CTkScrollableFrame behaves when there's no horizontal
        # overflow.
        content_req_width = self.content.winfo_reqwidth()
        self.canvas.itemconfigure(self._content_window, width=max(event.width, content_req_width))
        self._schedule_update()

    def _update_scrollbars(self):
        self.content.update_idletasks()
        self.canvas.update_idletasks()
        content_height = self.content.winfo_reqheight()
        content_width = self.content.winfo_reqwidth()
        viewport_height = self.canvas.winfo_height()
        viewport_width = self.canvas.winfo_width()

        needs_vertical = content_height > viewport_height
        needs_horizontal = content_width > viewport_width

        if needs_vertical:
            if not self.vbar.winfo_ismapped():
                # before=self.canvas is the actual fix -- without it, the
                # canvas (packed first, with expand=True) has already
                # claimed the entire cavity by the time this runs, so a
                # later, order-less pack() call for vbar gets 0x0 space
                # even though Tk reports it as successfully pack-managed.
                # hbar below already had this right; vbar didn't.
                self.vbar.pack(side="right", fill="y", before=self.canvas)
        else:
            if self.vbar.winfo_ismapped():
                self.vbar.pack_forget()

        if needs_horizontal:
            if not self.hbar.winfo_ismapped():
                self.hbar.pack(side="bottom", fill="x", before=self.canvas)
        else:
            if self.hbar.winfo_ismapped():
                self.hbar.pack_forget()

    def _bind_mousewheel(self):
        # Bound once, unconditionally, not on canvas <Enter>/<Leave> --
        # that only fired when hovering the thin strip of canvas not
        # covered by any content widget, which in practice was mostly
        # just the area right next to the scrollbar. Child widgets
        # embedded via create_window (every card, label, button packed
        # into self.content) have their own Enter/Leave semantics that
        # don't bubble up to the canvas, so hovering the actual content
        # never triggered this at all -- exactly matching "I have to
        # hover the scrollbar to activate."
        #
        # add="+" so multiple AdaptiveScrollFrame instances can coexist
        # without one's binding silently replacing another's; each
        # instance's handler independently checks whether the cursor is
        # actually over ITS OWN area before doing anything.
        #
        # Deliberately NOT unbound on <Destroy> -- unbind_all is global
        # (it clears every handler for that event sequence, not just
        # this instance's), so if two frames coexist and one is
        # destroyed first, cleaning up its own binding this way would
        # have silently broken the other, still-alive frame's scrolling
        # too. Instead, _on_mousewheel/_on_shift_mousewheel defensively
        # no-op if this instance's own widgets are already gone, which is
        # safe regardless of how many other instances exist.
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel, add="+")
        self.canvas.bind_all("<Shift-MouseWheel>", self._on_shift_mousewheel, add="+")

    def _cursor_is_over_this_frame(self, event):
        """
        Real bounds check, not a guess -- finds whatever widget is
        actually under the cursor right now (winfo_containing), then
        walks up its ancestor chain to see if this scroll frame's canvas
        is among them. This is what makes scrolling work no matter which
        child widget (card, label, button) the cursor happens to be over,
        instead of depending on Enter/Leave firing on the canvas itself.
        """
        try:
            widget = self.canvas.winfo_containing(event.x_root, event.y_root)
        except Exception:
            return False
        while widget is not None:
            if widget is self.canvas:
                return True
            try:
                widget = widget.master
            except Exception:
                return False
        return False

    def _on_mousewheel(self, event):
        try:
            if self.vbar.winfo_ismapped() and self._cursor_is_over_this_frame(event):
                self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except Exception:
            pass  # this instance's widgets no longer exist -- safe no-op, doesn't affect other instances

    def _on_shift_mousewheel(self, event):
        try:
            if self.hbar.winfo_ismapped() and self._cursor_is_over_this_frame(event):
                self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        except Exception:
            pass  # this instance's widgets no longer exist -- safe no-op, doesn't affect other instances

    def bind_keyboard_scroll(self, widget=None):
        """
        PgUp/PgDn/Home/End scrolling. Not bound automatically at
        construction time -- keyboard events need a specific widget with
        focus to bind against (usually the canvas itself, or the whole
        window for a screen with no other focusable controls competing for
        those keys). Call this explicitly once the caller knows which
        widget should own these keys, so this doesn't silently steal
        PgUp/PgDn from a text entry or table that also wants them.
        """
        target = widget or self.canvas
        target.bind("<Prior>", lambda e: self.canvas.yview_scroll(-1, "pages") if self.vbar.winfo_ismapped() else None)
        target.bind("<Next>", lambda e: self.canvas.yview_scroll(1, "pages") if self.vbar.winfo_ismapped() else None)
        target.bind("<Home>", lambda e: self.canvas.yview_moveto(0) if self.vbar.winfo_ismapped() else None)
        target.bind("<End>", lambda e: self.canvas.yview_moveto(1) if self.vbar.winfo_ismapped() else None)


class ScrollableDialog(ctk.CTkToplevel):
    """
    Base class for dialogs whose natural content height might exceed
    available screen space -- e.g. New Vendor's ~900px natural height on a
    laptop with limited usable screen height, which previously forced the
    user to manually resize the dialog just to reach Save/Cancel.

    Auto-sizes to fit within the actual screen (leaving margin for the
    taskbar and window chrome), and wraps content in AdaptiveScrollFrame so
    if the capped size still isn't enough, the FORM becomes scrollable
    rather than the dialog silently exceeding the screen.

    Subclasses build their fields into self.body (not self directly):

        class MyDialog(ScrollableDialog):
            def __init__(self, master, ...):
                super().__init__(master, title="My Dialog", natural_width=460, natural_height=700)
                ctk.CTkLabel(self.body, text="...").grid(...)
                ...

    Deliberately separate from being automatically applied to every large
    dialog in the app -- rolling this out to Material/BOQ/Invoice/Contract/
    Site Visit/Proposal/Package dialogs is real, planned future work, not
    done in this pass. This is applied to VendorForm only, the one
    concretely reported and confirmed broken; the pattern needs to prove
    itself there first.
    """
    def __init__(self, master, title, natural_width=460, natural_height=700):
        super().__init__(master)
        self.title(title)
        self.configure(fg_color=theme.PARCHMENT)
        self.transient(master.winfo_toplevel())
        self.grab_set()

        screen_height = self.winfo_screenheight()
        screen_width = self.winfo_screenwidth()
        # Leave real margin for taskbar + window title bar + breathing room
        # -- using the full screen height would still leave no room for
        # Windows' own chrome around the dialog.
        max_height = int(screen_height * 0.85)
        max_width = int(screen_width * 0.9)

        actual_width = min(natural_width, max_width)
        actual_height = min(natural_height, max_height)
        self.geometry(f"{actual_width}x{actual_height}")

        self.scroll_container = AdaptiveScrollFrame(self, fg_color=theme.PARCHMENT)
        self.scroll_container.pack(fill="both", expand=True)
        # Subclasses build their form fields into this, not self directly.
        self.body = self.scroll_container.content


class EntityWorkspace(ctk.CTkToplevel):
    """
    Generic, reusable workspace shell: breadcrumb -> header (avatar/name/
    tags/quick actions) -> tabs -> tab content. Built once so Contractor,
    Vendor, and any future entity workspace (Client, Employee, ...) share
    identical layout, spacing, and navigation instead of each being a
    separate, copy-pasted implementation -- a bug fix or enhancement here
    benefits every workspace built on it.

    Usage:
        ws = EntityWorkspace(master, breadcrumb_root="Contractors",
                             entity_name="Goutam Roy", tags=["Preferred", "Active"],
                             subtitle="Civil Work Contractor",
                             quick_actions=[("Edit", edit_fn), ("Archive", archive_fn)])
        ws.add_tab("Overview", build_overview_fn)
        ws.add_tab("Projects", build_projects_fn)
        ws.build()

    Each build_fn receives the tab's content frame (already scrollable via
    AdaptiveScrollFrame) and populates it. Tabs are built lazily -- only
    when first selected -- so a workspace with several tabs doesn't pay the
    cost of querying data for tabs the user never opens.
    """
    def __init__(self, master, breadcrumb_root, entity_name, subtitle=None, tags=None,
                quick_actions=None, on_close=None):
        super().__init__(master)
        self.title(entity_name)
        self.geometry("1050x720")
        self.minsize(900, 600)
        self.configure(fg_color=theme.PARCHMENT)
        self.transient(master.winfo_toplevel())
        self.protocol("WM_DELETE_WINDOW", self._on_close_wrapper(on_close))
        self.after(10, lambda: (self.lift(), self.focus_force()))

        self._tab_builders = {}
        self._built_tabs = set()

        # ---------------- Breadcrumb ----------------
        breadcrumb = ctk.CTkFrame(self, fg_color="transparent")
        breadcrumb.pack(fill="x", padx=20, pady=(15, 5))
        ctk.CTkLabel(breadcrumb, text=breadcrumb_root, font=theme.FONT_SMALL, text_color=theme.MUTED,
                    cursor="hand2").pack(side="left")
        ctk.CTkLabel(breadcrumb, text="  >  ", font=theme.FONT_SMALL, text_color=theme.MUTED).pack(side="left")
        ctk.CTkLabel(breadcrumb, text=entity_name, font=theme.FONT_SMALL, text_color=theme.INK).pack(side="left")

        # ---------------- Header ----------------
        header = ctk.CTkFrame(self, fg_color=theme.WHITE, corner_radius=8)
        header.pack(fill="x", padx=20, pady=(0, 10))
        header_left = ctk.CTkFrame(header, fg_color="transparent")
        header_left.pack(side="left", fill="both", expand=True, padx=15, pady=15)

        name_row = ctk.CTkFrame(header_left, fg_color="transparent")
        name_row.pack(anchor="w")
        ctk.CTkLabel(name_row, text=entity_name, font=theme.FONT_SUBHEADING, text_color=theme.INK).pack(side="left")
        for tag in (tags or []):
            ctk.CTkLabel(name_row, text=f"  {tag}  ", font=theme.FONT_SMALL, text_color=theme.WHITE,
                        fg_color=theme.BRASS, corner_radius=4).pack(side="left", padx=(8, 0))
        if subtitle:
            ctk.CTkLabel(header_left, text=subtitle, font=theme.FONT_BODY, text_color=theme.MUTED).pack(
                anchor="w", pady=(4, 0))

        if quick_actions:
            actions_frame = ctk.CTkFrame(header, fg_color="transparent")
            actions_frame.pack(side="right", padx=15, pady=15)
            for label, command in quick_actions:
                ctk.CTkButton(actions_frame, text=label, command=command, fg_color=theme.INK,
                             hover_color=theme.BRASS, font=theme.FONT_SMALL, height=30, width=110).pack(
                    side="left", padx=(6, 0))

        # ---------------- Tabs (added via add_tab(), built via build()) ----------------
        self.tabs = ctk.CTkTabview(self, fg_color=theme.WHITE, segmented_button_fg_color=theme.INK,
                                   segmented_button_selected_color=theme.BRASS,
                                   segmented_button_unselected_color=theme.INK,
                                   command=self._on_tab_change)
        self.tabs.pack(fill="both", expand=True, padx=20, pady=(0, 15))

    def _on_close_wrapper(self, on_close):
        def handler():
            if on_close:
                on_close()
            self.destroy()
        return handler

    def add_tab(self, name, build_fn):
        """Register a tab and its content-builder. Call build() once all tabs are added."""
        self._tab_builders[name] = build_fn

    def build(self):
        """Actually create the tabs and build the first one immediately (lazy for the rest)."""
        for name in self._tab_builders:
            self.tabs.add(name)
        first_tab = next(iter(self._tab_builders), None)
        if first_tab:
            self._build_tab_content(first_tab)

    def _on_tab_change(self):
        current = self.tabs.get()
        if current and current not in self._built_tabs:
            self._build_tab_content(current)

    def _build_tab_content(self, tab_name):
        scroll = AdaptiveScrollFrame(self.tabs.tab(tab_name), fg_color="transparent")
        scroll.pack(fill="both", expand=True)
        self._tab_builders[tab_name](scroll.content)
        self._built_tabs.add(tab_name)


def export_rows_to_csv(filepath, headers, rows):
    """headers: list of column names. rows: list of tuples/lists, same order as headers."""
    import csv
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def export_rows_to_excel(filepath, headers, rows, title="Report"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18
    wb.save(filepath)


def export_rows_to_pdf(filepath, title, headers, rows, metadata=None):
    """
    metadata: optional dict of label->value shown at the top (Generated On, Filters, etc.)

    Every exported report now carries a real letterhead (company name,
    tagline, address -- only confirmed-real business details, nothing
    fabricated: no invented website, phone number, or credentials that
    weren't actually confirmed), a repeating header/footer on every
    page (not just the first), expanded metadata (a real, deterministic
    Report ID and the app version), and a signature block at the end --
    the difference between a raw data export and something that reads
    as professionally prepared and identifiable if it's ever forwarded
    on its own. Uses the app's own established brand colors (Ink/Brass/
    Parchment) rather than a new palette that wouldn't match anything
    else ADS OS produces.
    """
    import datetime
    import hashlib
    import os
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Real bug found and fixed while testing this change, not a
    # cosmetic tweak: Helvetica (ReportLab's built-in base font, and
    # this function's previous default) has no Rupee glyph at all --
    # every currency amount in every exported report was silently
    # rendering as a black missing-glyph box, confirmed directly by
    # extracting text from a real generated PDF and seeing "included
    # 41,990" instead of "included 41,990" with a real symbol. DejaVu
    # Sans does include it, but Windows doesn't ship this font by
    # default, so it must be bundled with the app itself (fonts/
    # DejaVuSans*.ttf, loaded via a path relative to this script) --
    # referencing it by name alone would only work by accident on
    # whichever machine happens to already have it installed.
    fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
    regular_path = os.path.join(fonts_dir, "DejaVuSans.ttf")
    bold_path = os.path.join(fonts_dir, "DejaVuSans-Bold.ttf")
    if "ADSReport" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("ADSReport", regular_path))
        pdfmetrics.registerFont(TTFont("ADSReport-Bold", bold_path))
    FONT, FONT_BOLD = "ADSReport", "ADSReport-Bold"

    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    INK = colors.HexColor("#1A1A1A")
    BRASS = colors.HexColor("#B68100")
    MUTED = colors.HexColor("#6B6B6B")

    styles = getSampleStyleSheet()
    company_style = ParagraphStyle("Company", parent=styles["Normal"], fontName=FONT_BOLD,
                                   fontSize=14, textColor=INK, spaceAfter=2)
    tagline_style = ParagraphStyle("Tagline", parent=styles["Normal"], fontName=FONT,
                                   fontSize=8, textColor=BRASS, spaceAfter=2)
    address_style = ParagraphStyle("Address", parent=styles["Normal"], fontName=FONT,
                                   fontSize=8, textColor=MUTED)
    report_title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName=FONT_BOLD,
                                        fontSize=16, textColor=INK, spaceBefore=10, spaceAfter=4)
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontName=FONT, fontSize=8.5,
                                textColor=MUTED)
    signature_style = ParagraphStyle("Signature", parent=styles["Normal"], fontName=FONT,
                                     fontSize=9, textColor=INK)

    now = datetime.datetime.now()
    # Deterministic, real report ID -- derived from the actual title and
    # generation timestamp, not a random or fabricated value.
    report_id = f"RPT-{now.strftime('%Y%m%d')}-{hashlib.md5(f'{title}{now.isoformat()}'.encode()).hexdigest()[:6].upper()}"

    elements = [
        Paragraph("ATISH DIPANKAR'S CREATIVE SPACE", company_style),
        Paragraph("Architecture &nbsp;|&nbsp; Planning &nbsp;|&nbsp; Interior Design", tagline_style),
        Paragraph("Sky City Mall, Gangarampur, Dakshin Dinajpur, West Bengal", address_style),
        Spacer(1, 14),
        Paragraph(f"{title} Report" if not title.endswith("Report") else title, report_title_style),
    ]

    full_metadata = {
        "Generated On": now.strftime("%d %b %Y, %I:%M %p"),
        "Prepared By": "ADS Office Suite",
        "Report ID": report_id,
    }
    if metadata:
        full_metadata.update(metadata)
    for label, value in full_metadata.items():
        elements.append(Paragraph(f"<b>{label}:</b> {value}", meta_style))
    elements.append(Spacer(1, 12))

    table_data = [headers] + [list(row) for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F0E6")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    # Signature block -- once at the end of the document, not repeated
    # per page (unlike the footer below, which does repeat).
    elements.append(Spacer(1, 28))
    elements.append(Paragraph("Prepared By", meta_style))
    elements.append(Paragraph("Atish Dipankar's Creative Space", signature_style))
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("_" * 30, signature_style))
    elements.append(Paragraph("Authorized Signatory", meta_style))

    def _draw_page_furniture(canvas, doc):
        """Real header/footer drawn on every page, not only the first -- confidentiality line, company name, and a genuine page number."""
        canvas.saveState()
        page_width, page_height = landscape(A4)
        canvas.setStrokeColor(BRASS)
        canvas.setLineWidth(0.75)
        canvas.line(1.5 * cm, page_height - 1.1 * cm, page_width - 1.5 * cm, page_height - 1.1 * cm)
        canvas.line(1.5 * cm, 1.1 * cm, page_width - 1.5 * cm, 1.1 * cm)

        canvas.setFont(FONT, 8)
        canvas.setFillColor(MUTED)
        canvas.drawString(1.5 * cm, 0.7 * cm,
                          "Atish Dipankar's Creative Space -- Confidential, prepared for internal/client use only.")
        canvas.drawRightString(page_width - 1.5 * cm, 0.7 * cm, f"Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), topMargin=2.2 * cm, bottomMargin=1.6 * cm)
    doc.build(elements, onFirstPage=_draw_page_furniture, onLaterPages=_draw_page_furniture)


class ReportScreen(ctk.CTkFrame):
    """
    Generic, reusable report layout: breadcrumb -> filter bar -> summary
    cards -> data table -> export toolbar. Built once so every report
    shares identical structure, filter behavior, and export handling
    instead of each report reinventing all four.

    Subclasses (or callers) provide:
    - title: report name
    - build_data(start_date, end_date) -> (summary_cards, table_headers, table_rows)
      where summary_cards is a list of (value, label, sublabel) tuples,
      table_headers is a list of column names, table_rows is a list of
      tuples matching those columns.
    - on_row_open(row_index, row_data): optional drill-down callback, called
      on double-click of a table row.
    """
    def __init__(self, master, title, build_data, on_row_open=None, breadcrumb="Reports"):
        super().__init__(master, fg_color=theme.PARCHMENT)
        self.title_text = title
        self.build_data = build_data
        self.on_row_open = on_row_open
        self._current_rows = []
        self._current_headers = []
        self._build_ui(breadcrumb)
        # Opens unfiltered (all time), matching Financial Dashboard's own
        # default behavior -- NOT apply_filters(), which would read
        # DateEntry's default value (today) and silently filter out every
        # real record that isn't dated today, showing an empty/wrong report
        # on first open with no indication anything was filtered at all.
        self.reset_filters()

    def _build_ui(self, breadcrumb):
        crumb = ctk.CTkFrame(self, fg_color="transparent")
        crumb.pack(fill="x", padx=20, pady=(15, 0))
        ctk.CTkLabel(crumb, text=f"{breadcrumb}  >  {self.title_text}", font=theme.FONT_SMALL,
                    text_color=theme.MUTED).pack(anchor="w")

        ctk.CTkLabel(self, text=self.title_text, font=theme.FONT_HEADING, text_color=theme.INK).pack(
            anchor="w", padx=20, pady=(5, 10))

        filter_bar = ctk.CTkFrame(self, fg_color=theme.WHITE, corner_radius=8)
        filter_bar.pack(fill="x", padx=20, pady=(0, 10))
        ctk.CTkLabel(filter_bar, text="From", font=theme.FONT_SMALL, text_color=theme.MUTED).pack(
            side="left", padx=(15, 5), pady=10)
        from tkcalendar import DateEntry
        self.from_date = DateEntry(filter_bar, width=12, date_pattern="yyyy-mm-dd", background=theme.BRASS,
                                   foreground="white", borderwidth=1)
        self.from_date.pack(side="left", padx=(0, 15), pady=10)
        ctk.CTkLabel(filter_bar, text="To", font=theme.FONT_SMALL, text_color=theme.MUTED).pack(
            side="left", padx=(0, 5), pady=10)
        self.to_date = DateEntry(filter_bar, width=12, date_pattern="yyyy-mm-dd", background=theme.BRASS,
                                 foreground="white", borderwidth=1)
        self.to_date.pack(side="left", padx=(0, 15), pady=10)
        ctk.CTkButton(filter_bar, text="Apply", command=self.apply_filters, fg_color=theme.BRASS,
                     hover_color=theme.INK, font=theme.FONT_SMALL, width=80, height=28).pack(side="left", padx=(0, 8), pady=10)
        ctk.CTkButton(filter_bar, text="Reset (All Time)", command=self.reset_filters, fg_color="transparent",
                     hover_color=theme.PARCHMENT, text_color=theme.INK, font=theme.FONT_SMALL, height=28).pack(
            side="left", pady=10)

        export_frame = ctk.CTkFrame(filter_bar, fg_color="transparent")
        export_frame.pack(side="right", padx=15, pady=10)
        for label, fmt in [("CSV", "csv"), ("Excel", "xlsx"), ("PDF", "pdf")]:
            ctk.CTkButton(export_frame, text=f"Export {label}", command=lambda f=fmt: self._export(f),
                         fg_color=theme.INK, hover_color=theme.BRASS, font=theme.FONT_SMALL,
                         width=90, height=28).pack(side="left", padx=(4, 0))

        self.stats_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.stats_frame.pack(fill="x", padx=20, pady=(0, 10))

        table_frame = ctk.CTkFrame(self, fg_color=theme.WHITE)
        table_frame.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        self.tree_container = table_frame

    def reset_filters(self):
        import datetime
        self.from_date.set_date(datetime.date(2000, 1, 1))
        self.to_date.set_date(datetime.date.today())
        self.apply_filters(use_range=False)

    def apply_filters(self, use_range=True):
        start = self.from_date.get_date().isoformat() if use_range else None
        end = self.to_date.get_date().isoformat() if use_range else None
        summary_cards, headers, rows = self.build_data(start, end)
        self._current_headers = headers
        self._current_rows = rows

        for w in self.stats_frame.winfo_children():
            w.destroy()
        for value, label, sublabel in summary_cards:
            stat_card(self.stats_frame, value, label, sublabel).pack(side="left", padx=(0, 12))

        for w in self.tree_container.winfo_children():
            w.destroy()
        if not rows:
            empty_state(self.tree_container, "No data for this date range.")
            return
        cols = tuple(range(len(headers)))
        tree = ttk.Treeview(self.tree_container, columns=cols, show="headings", height=16)
        for i, h in enumerate(headers):
            tree.heading(i, text=h)
            tree.column(i, width=120)
        for row_idx, row in enumerate(rows):
            tree.insert("", "end", iid=str(row_idx), values=row)
        if self.on_row_open:
            tree.bind("<Double-1>", lambda e: self._handle_row_open(tree))
        vscroll = ttk.Scrollbar(self.tree_container, orient="vertical", command=tree.yview)
        tree.configure(yscroll=vscroll.set)
        vscroll.pack(side="right", fill="y")
        hscroll = ttk.Scrollbar(self.tree_container, orient="horizontal", command=tree.xview)
        tree.configure(xscroll=hscroll.set)
        hscroll.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True, side="left")

    def _handle_row_open(self, tree):
        sel = tree.selection()
        if sel:
            self.on_row_open(int(sel[0]), self._current_rows[int(sel[0])])

    def _export(self, fmt):
        import datetime
        from tkinter import filedialog
        default_name = f"{self.title_text.replace(' ', '_')}_{datetime.date.today().isoformat()}.{fmt}"
        filepath = filedialog.asksaveasfilename(defaultextension=f".{fmt}", initialfile=default_name,
                                                filetypes=[(fmt.upper(), f"*.{fmt}")])
        if not filepath:
            return
        try:
            if fmt == "csv":
                export_rows_to_csv(filepath, self._current_headers, self._current_rows)
            elif fmt == "xlsx":
                export_rows_to_excel(filepath, self._current_headers, self._current_rows, title=self.title_text)
            elif fmt == "pdf":
                metadata = {
                    "Generated On": datetime.datetime.now().strftime("%d %b %Y, %I:%M %p"),
                    "Date Range": f"{self.from_date.get_date().isoformat()} to {self.to_date.get_date().isoformat()}",
                }
                export_rows_to_pdf(filepath, self.title_text, self._current_headers, self._current_rows, metadata)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Export failed", f"Could not export: {e}", parent=self)
